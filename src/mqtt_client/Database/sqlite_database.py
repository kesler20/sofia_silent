import time
import datetime
import pandas as pd
from pandas import DataFrame
import openpyxl as xl
import sqlite3
import xlrd
from logs_src.main_logging import logger

class File(object):
    
    def __init__(self, filename, mode):
        self.filename = filename
        self.mode = mode

    def __enter__(self):
        self.file = open(self.filename, self.mode)
        return self.file

    def __exit__(self, exec_type, exec_val, traceback):
        self.file.close()

# Textual month, day and year
today = datetime.date.today().weekday()

def get_month_name():
    month_num = str(time.strftime("%m"))
    datetime_object = datetime.datetime.strptime(month_num, "%m")
    return datetime_object.strftime("%B")


def contruct_column_of_strings(number_of_columns, sentence):
    with File(r'Database\code.txt', 'w') as w:
        for i in range(number_of_columns):
            w.write(sentence)
    with File(r'Database\code.txt', 'r') as r:
        content = r.read()

def return_sheetname(filename):
    wb = xl.load_workbook(filename)
    return wb.sheetnames


def remove_excel_sheet(filename, sheetname):
    wb = xl.load_workbook(filename)
    xl_sheet = wb.get_sheet_by_name(sheetname)
    wb.remove_sheet(xl_sheet)
    wb.save(filename)


def append_to_excel(df: DataFrame, sheetname, filename, copy_old=False, spaces=1):
    try:
        old_df = pd.read_excel(filename, sheetname, engine='openpyxl')
    except xlrd.biffh.XLRDError:
        wb = xl.load_workbook(filename)
        wb.create_sheet(sheetname)
        wb.save(filename)

    if copy_old:
        old_df = pd.read_excel(filename, sheetname, engine='openpyxl')
        collection_of_dfs = [old_df]
        collection_of_dfs.append(df)
    else:
        collection_of_dfs = [df]

    remove_excel_sheet(filename, sheetname)

    writer = pd.ExcelWriter(r'Database\assets\data.xlsx', engine='xlsxwriter')
    row = 0
    for dataframe in collection_of_dfs:
        dataframe.to_excel(writer, sheet_name='DATA',
                           startrow=row, startcol=0, index=False)
        row = row + len(dataframe.index) + spaces + 1
    writer.save()

    updated_df = pd.read_excel(r'Database\assets\data.xlsx', engine='openpyxl')

    logger.info(updated_df)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as f:
        updated_df.to_excel(f, sheet_name=sheetname, index=False)
        f.save()


def get_all_tables_in_database(database_filename=r'Database\mydata.db'):
    try:

        # Making a connection between sqlite3
        # database and Python Program
        sqliteConnection = sqlite3.connect(database_filename, 15)

        # If sqlite3 makes a connection with python
        # program then it will logger.info "Connected to SQLite"
        # Otherwise it will show errors
        logger.info("Connected to SQLite")

        # Getting all tables from sqlite_master
        sql_query = """SELECT name FROM sqlite_master
        WHERE type='table';"""

        # Creating cursor object using connection object
        cursor = sqliteConnection.cursor()

        # executing our sql query
        cursor.execute(sql_query)
        logger.info("List of tables\n")

        # logger.infoing all tables list
        return cursor.fetchall()

    except sqlite3.Error as error:
        logger.info("Failed to execute the above query", error)

    finally:

        # Inside Finally Block, If connection is
        # open, we need to close it
        if sqliteConnection:

            # using close() method, we will close
            # the connection
            sqliteConnection.close()

            # After closing connection object, we
            # will logger.info "the sqlite connection is
            # closed"
            logger.info("the sqlite connection is closed")

def contruct_str_metadata_column(number_of_columns, sentence):
    with File(r'Database\sql_command.txt', 'w') as w:
        for i in range(number_of_columns):
            if i == 0:
                w.write(f'entry{i} : {sentence[i]}')
            else:
                w.write(f'            entry{i} : {sentence[i]}')
    with File(r'Database\sql_command.txt', 'r') as r:
        content = r.read()
    return content


def column_of_entries(number_of_columns, entries):
    with File(r'Database\sql_command1.txt', 'w') as w:
        for i in range(number_of_columns):
            if i == 0:
                w.write(f'entry{i} : {entries[i]}\n')
            else:
                w.write(f'            entry{i} : {entries[i]}\n')
    with File(r'Database\sql_command1.txt', 'r') as r:
        content = r.read()
    return content


def rows_of_entries(number_of_rows, entries):
    with File(r'Database\sql_command2.txt', 'w') as w:
        for i in range(number_of_rows):
            if i == number_of_rows - 1:
                if type(entries[i]) == str:
                    w.write(f"'{entries[i]}' ")
                else:
                    w.write(f"{entries[i]} ")
            else:
                if type(entries[i]) == str:
                    w.write(f"'{entries[i]}', ")
                else:
                    w.write(f"{entries[i]}, ")

    with File(r'Database\sql_command2.txt', 'r') as r:
        content = r.read()
    return content


def create_table_column_of_entries(number_of_columns, entries):
    with File(r'Database\sql_command3.txt', 'w') as w:
        for i in range(number_of_columns):
            if type(entries[i]) == str:
                sql_type = 'TEXT'
            elif type(entries[i]) == int or float:
                sql_type = 'INTEGER'
            else:
                sql_type = 'TEXT'
            if i == 0:
                w.write(f'entry{i} {sql_type},\n')
            elif i == number_of_columns - 1:
                w.write(f'            entry{i} {sql_type},')
            else:
                w.write(f'            entry{i} {sql_type},\n')
    with File(r'Database\sql_command3.txt', 'r') as r:
        content = r.read()
    return content


def create_table_column_of_entries_for_child(number_of_columns, entries):
    with File(r'Database\sql_command4.txt', 'w') as w:
        for i in range(number_of_columns):
            if type(entries[i]) == str:
                sql_type = 'TEXT'
            elif type(entries[i]) == int:
                sql_type = 'INTEGER'
            if i == 0:
                w.write(f'entry{i} {sql_type},\n')
            elif i == number_of_columns - 1:
                w.write(f'            entry{i} {sql_type}')
            else:
                w.write(f'            entry{i} {sql_type},\n')
    with File(r'Database\sql_command4.txt', 'r') as r:
        content = r.read()
    return content

class Entity:

    def __init__(self, *args):
        self.entries = [arg for arg in args[3:]]
        self.name = args[0]
        self.hirerarchy = args[1] if len(self.entries) >= 2 else 'PARENT'
        self.parent = args[2] if len(self.entries) >= 3 else 'PARENT'
        self._id = self.increment_id()
        self.date_id = str(time.strftime("%d/%m/%Y"))
        self.saved = False
        # remember to #commit and close the connection at the end of each action

    def __str__(self):
        content = contruct_str_metadata_column(
            len(self.entries), [str(entry) + '\n' for entry in self.entries])
        return f'''
        {self.name}(
            saved : {self.saved}
            ID : {self._id}
            {content}
        )
        '''

    def increment_id(self):
        self.create_all()
        connection = sqlite3.connect(r'Database\mydata.db', 15)
        cursor = connection.cursor()
        cursor.execute(f'''
        SELECT * FROM {self.name}
        ''')
        rows = cursor.fetchall()
        connection.commit()
        connection.close()
        _id = len(rows) + 1
        logger.info(_id)
        return _id

    def add_entries(self):
        self.create_all()
        content = rows_of_entries(len(self.entries), self.entries)
        logger.info('add entries content', content)
        if self.hirerarchy == 'CHILD':
            command = f'''
        INSERT INTO {self.name} VALUES
        ('{self.date_id}',{content})
        '''
            logger.info(command)
            self.cursor.execute(command)
        else:
            command = f'''
        INSERT INTO {self.name} VALUES
        ({self._id}, {content},'{self.date_id}')
        '''
            logger.info(command)
            self.cursor.execute(command)
        self.saved = True
        self.connection.commit()
        self.connection.close()

    def load_entry(self, _id):
        self.create_all()
        try:
            self.cursor.execute(f'''
        SELECT * FROM {self.name}
        WHERE ID = {_id}
            ''')
        except sqlite3.OperationalError:
            pass

        Entity_ = self.cursor.fetchone()
        self.connection.close()
        return Entity_

    def create_all(self):
        database_name = r'Database\mydata.db'
        connection = sqlite3.connect(database_name, 15)
        self.cursor = connection.cursor()
        content = create_table_column_of_entries(
            len(self.entries), self.entries)
        command = f'''
        CREATE TABLE IF NOT EXISTS {self.name} (
            ID INTEGER PRIMARY KEY,
            {content}
            date_id TEXT NOT NULL,
            FOREIGN KEY (date_id)
            REFERENCES children (date_id) 
        );
        '''
        logger.info(command)

        self.cursor.execute(command)

        if self.hirerarchy == 'CHILD':
            # drop the parent entity which is created by  default and create the parent entity using the known name
            command = f'''
            DROP TABLE {self.name};
            '''
            logger.info(command, ' because it is a child')
            self.cursor.execute(command)

            command = f'''
        CREATE TABLE IF NOT EXISTS {self.parent} (
            ID INTEGER PRIMARY KEY,
            {content}
            date_id TEXT NOT NULL,
            FOREIGN KEY (date_id)
            REFERENCES children (date_id) 
        );
        '''
            logger.info(command)
            self.cursor.execute(command)
            content = create_table_column_of_entries_for_child(
                len(self.entries), self.entries)
            command = f'''
        CREATE TABLE IF NOT EXISTS {self.name} (
            date_id TEXT PRIMARY KEY,
            {content}
        );
        '''
            logger.info(command)
            self.cursor.execute(command)

        connection.commit()
        self.connection = connection

#day = Entity('Monday','PARENT','',1,1,2,'this is so cool')
# day.add_entries()
#exercises = Entity('Triceps_extensions','CHILD','Monday',20,20,0)
# exercises.add_entries()
# logger.info(day)
# logger.info(exercises)
#entity = Entity('Monday')
#historical_day = entity.load_entry(1)
# logger.info(historical_day)
